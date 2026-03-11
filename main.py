import os
import re
import warnings
import pandas as pd
import unicodedata
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import streamlit as st


warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    module="openpyxl"
)


def hacer_encabezados_unicos(headers):
    usados = {}
    resultado = []

    for i, h in enumerate(headers, start=1):
        if pd.isna(h) or str(h).strip() == "":
            base = f"__col_{i}"
        else:
            base = str(h).strip()

        usados[base] = usados.get(base, 0) + 1

        if usados[base] == 1:
            resultado.append(base)
        else:
            resultado.append(f"{base} [{usados[base]}]")

    return resultado


def normalizar_texto(valor):
    if pd.isna(valor):
        return ""

    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)

    texto = str(valor).strip()

    if texto.lower() in {"", "nan", "none", "nat", "<na>"}:
        return ""

    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = texto.lower()
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()

    return texto


def construir_mapa_columnas_excel(encabezados):
    return {
        col: f"{get_column_letter(i + 1)} | {col}"
        for i, col in enumerate(encabezados)
    }


def leer_excel_para_preview(archivo_bytes, nombre_hoja, fila_encabezados_excel):
    df_raw = pd.read_excel(
        BytesIO(archivo_bytes),
        sheet_name=nombre_hoja,
        header=None,
        dtype=object
    )

    if fila_encabezados_excel < 1:
        raise ValueError("La fila de encabezados debe ser 1 o mayor.")

    if fila_encabezados_excel > len(df_raw):
        raise ValueError("La fila de encabezados está fuera del rango del archivo.")

    encabezados_originales = df_raw.iloc[fila_encabezados_excel - 1].tolist()
    encabezados = hacer_encabezados_unicos(encabezados_originales)

    df = df_raw.iloc[fila_encabezados_excel:].copy()
    df.columns = encabezados
    df.reset_index(drop=True, inplace=True)

    return df_raw, df, encabezados


def validar_y_generar_excel(
    archivo_bytes,
    nombre_archivo,
    nombre_hoja,
    fila_encabezados_excel,
    encabezados,
    df,
    grupo_a,
    grupo_b,
):
    wb = load_workbook(
        BytesIO(archivo_bytes),
        keep_vba=nombre_archivo.lower().endswith(".xlsm")
    )
    ws = wb[nombre_hoja]

    relleno_verde = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
    relleno_rojo = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")

    conteo_verdes = 0
    conteo_rojos = 0
    conteo_vacios = 0

    mapa_col_excel = {col: idx + 1 for idx, col in enumerate(encabezados)}

    for i, fila in df.iterrows():
        fila_excel = fila_encabezados_excel + 1 + i

        textos_a = [normalizar_texto(fila[col]) for col in grupo_a]
        textos_a = [t for t in textos_a if t != ""]

        for col_b in grupo_b:
            valor_b = fila[col_b]
            texto_b = normalizar_texto(valor_b)

            col_excel = mapa_col_excel[col_b]
            celda = ws.cell(row=fila_excel, column=col_excel)

            if texto_b == "":
                conteo_vacios += 1
                continue

            encontrado = any(texto_b in texto_a_normalizado for texto_a_normalizado in textos_a)

            if encontrado:
                celda.fill = relleno_verde
                conteo_verdes += 1
            else:
                celda.fill = relleno_rojo
                conteo_rojos += 1

    salida = BytesIO()
    wb.save(salida)
    salida.seek(0)

    nombre_base, extension = os.path.splitext(nombre_archivo)
    nombre_salida = f"{nombre_base}_validado{extension}"

    resumen = {
        "verdes": conteo_verdes,
        "rojos": conteo_rojos,
        "vacios": conteo_vacios,
        "nombre_salida": nombre_salida,
    }

    return salida, resumen


st.set_page_config(page_title="Validación de datos en Excel", layout="wide")
st.title("Validación de datos en Excel")

st.write(
    "Sube tu archivo, elige la hoja, indica la fila de encabezados "
    "y selecciona los grupos de columnas."
)

archivo = st.file_uploader(
    "Sube tu archivo Excel",
    type=["xlsx", "xlsm"],
    key="archivo_excel"
)

if archivo is not None:
    archivo_bytes = archivo.read()

    try:
        xls = pd.ExcelFile(BytesIO(archivo_bytes))
        hojas = xls.sheet_names

        col1, col2 = st.columns([2, 1])

        with col1:
            nombre_hoja = st.selectbox(
                "Selecciona la hoja",
                hojas,
                key="nombre_hoja"
            )

        with col2:
            fila_encabezados_excel = st.number_input(
                "Fila de encabezados",
                min_value=1,
                value=1,
                step=1,
                key="fila_encabezados_excel"
            )

        if st.button("Cargar vista previa", width="stretch"):
            st.session_state["cargar_preview"] = True

        if st.session_state.get("cargar_preview", False):
            df_raw, df, encabezados = leer_excel_para_preview(
                archivo_bytes=archivo_bytes,
                nombre_hoja=nombre_hoja,
                fila_encabezados_excel=int(fila_encabezados_excel),
            )

            mapa_columnas_excel = construir_mapa_columnas_excel(encabezados)

            st.subheader("Encabezados detectados")
            encabezados_mostrados = pd.DataFrame(
                {
                    "Columna Excel": [get_column_letter(i + 1) for i in range(len(encabezados))],
                    "Encabezado": encabezados
                }
            )
            st.dataframe(encabezados_mostrados, width="stretch", hide_index=True)

            st.subheader("Primeras 5 filas")
            df_preview = df.loc[~df.isna().all(axis=1)].head(5).copy()
            df_preview_mostrar = df_preview.fillna("").astype(str)
            st.dataframe(df_preview_mostrar, width="stretch")

            col_a, col_b = st.columns(2)

            with col_a:
                grupo_a = st.multiselect(
                    "Grupo A (donde buscará)",
                    options=encabezados,
                    format_func=lambda col: mapa_columnas_excel[col],
                    key="grupo_a"
                )

            with col_b:
                grupo_b = st.multiselect(
                    "Grupo B (datos a validar)",
                    options=encabezados,
                    format_func=lambda col: mapa_columnas_excel[col],
                    key="grupo_b"
                )

            if st.button("Validar y generar archivo", width="stretch"):
                if not grupo_a:
                    st.warning("Selecciona al menos una columna en el Grupo A.")
                elif not grupo_b:
                    st.warning("Selecciona al menos una columna en el Grupo B.")
                else:
                    salida, resumen = validar_y_generar_excel(
                        archivo_bytes=archivo_bytes,
                        nombre_archivo=archivo.name,
                        nombre_hoja=nombre_hoja,
                        fila_encabezados_excel=int(fila_encabezados_excel),
                        encabezados=encabezados,
                        df=df,
                        grupo_a=grupo_a,
                        grupo_b=grupo_b,
                    )

                    st.success("Proceso terminado.")
                    st.write(f"Celdas verdes: {resumen['verdes']}")
                    st.write(f"Celdas rojas: {resumen['rojos']}")
                    st.write(f"Celdas vacías en grupo B: {resumen['vacios']}")

                    st.download_button(
                        label="Descargar Excel validado",
                        data=salida.getvalue(),
                        file_name=resumen["nombre_salida"],
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        width="stretch",
                    )

    except Exception as e:
        st.error(f"No se pudo procesar el archivo.\n\n{e}")